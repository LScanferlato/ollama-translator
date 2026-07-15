#!/usr/bin/env python3
"""
Traduttore multilingua per CSV / XLSX / XLS.
Utilizza Ollama (LLM locale) per tradurre termini tecnici.

Uso:
    python translate_csv.py file.csv
    python translate_csv.py file.xlsx
    python translate_csv.py file.xls

Il file viene modificato direttamente (in-place) per consentire
arresto e ripresa senza perdere progressi.
"""

import argparse
import csv
import os
import pickle
import re
import select
import sys
import tempfile
import termios
import threading
import time
import tty
from pathlib import Path

import requests

OLLAMA_BASE = "http://localhost:11434"
OLLAMA_URL = OLLAMA_BASE + "/api/generate"
OLLAMA_TAGS_URL = OLLAMA_BASE + "/api/tags"
MODEL = "abb-decide/apertus-tools:8b-instruct-2509-q4_k_m"


def normalizza_base(url):
    """Assicura che l'URL base abbia lo schema http(s):// e niente slash
    finale. Accetta anche indirizzi senza schema (es. '192.168.1.5:11434')."""
    u = (url or "").strip()
    if not u:
        return u
    if not u.startswith(("http://", "https://")):
        u = "http://" + u
    return u.rstrip("/")


class Ritmo:
    """Delay adattivo per server: diminuisce (AIMD) quando il server
    risponde senza problemi, aumenta (backoff) in caso di errori/retry,
    così l'esecuzione viene ottimizzata senza sovraccaricare."""

    MIN = 0.05
    MAX = 10.0
    DECREASE = 0.1   # moltiplicativo, su successo
    INCREASE = 0.1    # additivo, su congestione/errore

    def __init__(self, start):
        self.delay = max(self.MIN, float(start))

    def ok(self):
        self.delay = max(self.MIN, self.delay * self.DECREASE)

    def congestione(self):
        self.delay = min(self.MAX, self.delay + self.INCREASE)
DELAY = 0.5
LINGUA_PARTENZA = "italiano"

COLONNA_SOURCE = "IT"
ESCLUSE = {"ID", "Categoria", "Termine", "Descrizione", "IT"}

# ─── Glossario ───────────────────────────────────────────────────
# File glossario multilingua (es. EU IATE). Contiene la colonna
# "source_term" (spesso in inglese) e la colonna "it" (italiano),
# oltre a molte altre lingue. Viene usato per completare le lingue
# mancanti nei file da tradurre e come riferimento per l'LLM.
GLOSSARIO_DEFAULT = "glossary_multilingual.csv"
GLOSSARIO_CACHE = "glossary_index.pkl"

# Mappa i codici lingua usati nei file da tradurre alle colonne del
# glossario. Le lingue non presenti nel glossario (es. EN) restano
# affidate all'LLM.
MAPPA_GLOSSARIO = {
    "DE": "de", "FR": "fr", "ES": "es", "PT": "pt", "RU": "ru",
    "ZH": "zh-Hans", "JA": "ja", "KO": "ko", "AR": "ar", "NL": "nl",
    "PL": "pl", "TR": "tr", "CS": "cs", "DA": "da", "FI": "fi",
    "EL": "el", "HU": "hu", "RO": "ro", "SV": "sv", "BG": "bg",
    "HR": "hr", "SK": "sk", "SL": "sl", "LT": "lt", "LV": "lv",
    "ET": "et", "HE": "he", "HI": "hi", "ID": "id", "MS": "ms",
    "TH": "th", "VI": "vi", "UK": "uk", "SR": "sr-Latn", "NO": "nb-NO",
}

TUTTE_LINGUE = ["EN", "DE", "FR", "ES", "PT", "RU", "ZH", "JA", "KO",
                "AR", "NL", "PL", "TR", "CS", "DA", "FI", "EL", "HU",
                "RO", "SV", "BG", "HR", "SK", "SL", "LT", "LV", "ET",
                "HE", "HI", "ID", "MS", "TH", "VI", "UK", "SR", "NO"]

MAX_RETRIES = 3
RETRY_DELAY = 2
SALVA_OGNI_N = 0

DEBUG = True


log_lock = threading.Lock()


def log(msg):
    with log_lock:
        print(msg, flush=True)


# ─── Colori per lingua ─────────────────────────────────────────
COLORI = [
    "\033[31m", "\033[32m", "\033[33m", "\033[34m", "\033[35m", "\033[36m",
    "\033[91m", "\033[92m", "\033[93m", "\033[94m", "\033[95m", "\033[96m",
]
RESET = "\033[0m"


def colore_lingua(code):
    if not code or not sys.stdout.isatty():
        return code
    h = abs(hash(code)) % len(COLORI)
    return f"{COLORI[h]}{code}{RESET}"


# ─── Interruzione con ESC ────────────────────────────────────────
interruzione = threading.Event()


def _ascolta_esc():
    if not sys.stdin.isatty():
        return
    fd = sys.stdin.fileno()
    try:
        old = termios.tcgetattr(fd)
    except Exception:
        return
    try:
        tty.setraw(fd)
        while not interruzione.is_set():
            r, _, _ = select.select([fd], [], [], 0.2)
            if r:
                try:
                    ch = sys.stdin.read(1)
                except Exception:
                    break
                if ch == "\x1b":  # ESC
                    interruzione.set()
                    break
    finally:
        try:
            termios.tcsetattr(fd, termios.TCSADRAIN, old)
        except Exception:
            pass


def avvia_controllo_esc():
    if not sys.stdin.isatty():
        return None
    t = threading.Thread(target=_ascolta_esc, daemon=True)
    t.start()
    return t


def ferma_controllo_esc(t):
    interruzione.set()
    if t and t.is_alive():
        t.join()


def norm(s):
    """Normalizza un termine per il confronto: minuscolo, spazi ridotti,
    virgolette e punteggiatura perimetrale rimosse."""
    if not s:
        return ""
    s = str(s)
    for q in ['"', "'", "«", "»", "“", "”", "‘", "’"]:
        s = s.strip().strip(q)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def prima_variante(s):
    """Le celle del glossario possono contenere più varianti separate da
    ' | '. Restituisce solo la prima, più pulita."""
    if not s:
        return ""
    s = s.strip()
    if " | " in s:
        s = s.split(" | ")[0].strip()
    for q in ['"', "'", "«", "»", "“", "”", "‘", "’"]:
        s = s.strip().strip(q)
    return s.strip()


def nome_lingua(codice):
    mappa = {
        "EN": "inglese", "DE": "tedesco", "FR": "francese",
        "ES": "spagnolo", "PT": "portoghese", "RU": "russo",
        "ZH": "cinese", "JA": "giapponese", "KO": "coreano",
        "AR": "arabo", "NL": "olandese", "PL": "polacco",
        "TR": "turco", "CS": "ceco", "DA": "danese",
        "FI": "finlandese", "EL": "greco", "HU": "ungherese",
        "RO": "rumeno", "SV": "svedese", "BG": "bulgaro",
        "HR": "croato", "SK": "slovacco", "SL": "sloveno",
        "LT": "lituano", "LV": "lettone", "ET": "estone",
        "HE": "ebraico", "HI": "hindi", "ID": "indonesiano",
        "MS": "malese", "TH": "thailandese", "VI": "vietnamita",
        "UK": "ucraino", "SR": "serbo", "NO": "norvegese",
    }
    return mappa.get(codice.upper(), codice)


def pulisci_risposta(testo, lingua):
    if not testo:
        return testo

    # toglie tutto dopo un doppio a capo
    testo = testo.split('\n\n')[0].split('\n')[0]

    # toglie prefissi come "Raspunsul este:", "The translation is:", ecc.
    testo = re.sub(r'(?i)^.*?(Răspunsul|Raspunsul|Raspunziul|Răspuns|Raspuns|The translation|Translation|In response|Nota|Note|Ответ|答案|Übersetzung|Traduction|Traduzione|翻訳|번역)[^:]*:\s*', '', testo)

    # se la risposta contiene ancora parti in italiano o sembra una spiegazione, estrai la traduzione
    parole = testo.split()
    if len(parole) > 5:
        # prova a estrarre da virgolette (ascii e unicode)
        virgolette = re.findall(r'["""\u201c\u201d\u2018\u2019]([^""\u201c\u201d\u2018\u2019]*)["""\u201c\u201d\u2018\u2019]', testo)
        if virgolette:
            testo = virgolette[-1]
        else:
            # prova a prendere solo l'ultima parte dopo "è" / "este" / "is" / "se traduce come"
            ultimo = re.split(r'(?i)\b(è|este|is|se traduce come|si traduce come|means|significa|est| heißt)\b', testo)
            if len(ultimo) > 1:
                testo = ultimo[-1].strip().strip('.').strip('"').strip("'").strip('»').strip('«')

    # toglie "Translated from X to Y: Z" o "Traduzione da X a Y: Z"
    testo = re.sub(r'(?i)(translated|traduzion?e?|traduci)\s*(da|from).+?(a|to)\s*\S+\s*[:|]\s*', '', testo)

    # toglie spiegazioni come "X in Y si traduce come Z"
    testo = re.sub(r'(?i).+?\b(si traduce|se traduce|significa|means|translates? to|in\s+\S+\s+is)\s+["“]?', '', testo)
    testo = re.sub(r'(?i).+?\b(in|su|en)\s+\S+\s+(si dice|se dice|è)\s+["“]?', '', testo)

    # toglie spiegazioni in parentesi (es. "이포드 기어 (Iphord Gear)")
    testo = re.sub(r'\s*\([^)]*\)\s*', '', testo)

    # se contiene =, prende la parte dopo l'ultimo =
    if '=' in testo:
        testo = testo.rsplit('=', 1)[-1]

    # se contiene : con frase corta dopo, prende quella
    if ':' in testo:
        parti = testo.rsplit(':', 1)
        if len(parti[-1].strip().split()) <= 8:
            testo = parti[-1]

    # rimuove virgolette di ogni tipo (anche unicode) da inizio e fine
    testo = testo.strip()
    for q in ['"', "'", '«', '»', '\u201c', '\u201d', '\u2018', '\u2019', '\u201e', '\u201a', '\u00ab', '\u00bb']:
        if testo.startswith(q):
            testo = testo[len(q):]
        if testo.endswith(q):
            testo = testo[:-len(q)]
    testo = testo.strip('. ')

    return testo


def traduci(testo, lingua, modello=MODEL, session=None, contesto=None, base_url=OLLAMA_URL, ritmo=None):
    lingua_nome = nome_lingua(lingua)
    prompt = (
        f"{testo}\n\n"
        f"Traduci il termine tecnico qui sopra in {lingua_nome}. "
    )
    if contesto:
        riferimenti = "; ".join(
            f"{nome_lingua(c)}: {v}" for c, v in contesto.items() if v and str(v).strip()
        )
        if riferimenti:
            prompt += (
                f"Usa come riferimento queste traduzioni nelle altre lingue "
                f"per coerenza: {riferimenti}. "
            )
    prompt += (
        f"Rispondi con una sola parola o frase breve. "
        f"Niente spiegazioni, note, prefissi o testo aggiuntivo."
    )
    payload = {
        "model": modello,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.1},
    }
    sess = session or requests
    for tentativo in range(1, MAX_RETRIES + 1):
        try:
            resp = sess.post(base_url, json=payload, timeout=120)
            resp.raise_for_status()
            result = resp.json()
            raw = result["response"].strip()
            if ritmo is not None:
                if tentativo > 1:
                    ritmo.congestione()
                else:
                    ritmo.ok()
            return pulisci_risposta(raw, lingua)
        except Exception as e:
            if tentativo < MAX_RETRIES:
                attesa = RETRY_DELAY * tentativo
                log(f"  ERRORE (tentativo {tentativo}/{MAX_RETRIES}): {e} -> riprovo tra {attesa}s")
                time.sleep(attesa)
                if ritmo is not None:
                    ritmo.congestione()
            else:
                log(f"  ERRORE traduzione '{testo}' -> {lingua}: {e}")
                if ritmo is not None:
                    ritmo.congestione()
                return ""
    return ""


def traduci_se_pieno(testo, lingua, session):
    if not testo or not testo.strip():
        return ""
    return traduci(testo, lingua, session=session)


# ─── Glossario ───────────────────────────────────────────────────

def carica_glossario(path, lingue_target, force_rebuild=False):
    """Carica (o costruisce/cacheizza) l'indice del glossario.

    Restituisce (indice, gloss_cols) dove:
      - indice: dict termine_normalizzato -> numero di record (0-based)
                indicizzato sulle colonne 'source_term', 'it' e sulle
                colonne del glossario corrispondenti alle lingue target.
      - gloss_cols: dict codice_lingua_target -> nome colonna glossario.
    """
    cache_path = Path(GLOSSARIO_CACHE)
    p = Path(path)

    if not force_rebuild and cache_path.exists():
        try:
            with open(cache_path, "rb") as f:
                blob = pickle.load(f)
            if blob.get("source") == str(p) and blob.get("size") == p.stat().st_size:
                log(f"Glossario caricato dalla cache: {cache_path.name}")
                return blob["indice"], blob["gloss_cols"]
            log("Cache glossario non aggiornata, la ricostruisco...")
        except Exception as e:
            log(f"Cache glossario illeggibile ({e}), la ricostruisco...")

    with open(p, newline="", encoding="utf-8-sig") as f:
        header = [h.strip() for h in next(csv.reader(f))]

    gloss_cols = {}
    for code in lingue_target:
        gcol = MAPPA_GLOSSARIO.get(code)
        if gcol and gcol in header:
            gloss_cols[code] = gcol

    index_cols = ["source_term", "it"] + list(gloss_cols.values())
    index_cols = [c for c in index_cols if c in header]

    indice = {}
    with open(p, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader):
            for c in index_cols:
                v = row.get(c, "")
                if v and v.strip():
                    k = norm(v)
                    if k and k not in indice:
                        indice[k] = idx

    try:
        with open(cache_path, "wb") as f:
            pickle.dump(
                {"source": str(p), "size": p.stat().st_size,
                 "indice": indice, "gloss_cols": gloss_cols},
                f,
            )
        log(f"Cache glossario salvata: {cache_path.name} ({len(indice)} termini)")
    except Exception as e:
        log(f"Impossibile salvare la cache del glossario: {e}")

    return indice, gloss_cols


def estrai_righe_glossario(path, record_ids, gloss_cols):
    """Estrae da disco solo le righe del glossario indicizzate in
    record_ids, restituendo dict record_id -> {colonna_glossario: valore}."""
    ris = {}
    if not record_ids:
        return ris
    wanted = set(record_ids)
    cols = list(gloss_cols.values())
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader):
            if idx in wanted:
                d = {}
                for gcol in cols:
                    v = row.get(gcol, "")
                    if v and v.strip():
                        d[gcol] = v
                ris[idx] = d
                if len(ris) == len(wanted):
                    break
    return ris


# ─── CSV ───────────────────────────────────────────────────────────

def rileva_delimitatore(prima_riga):
    virgole = prima_riga.count(",")
    punti = prima_riga.count(";")
    return ";" if punti >= virgole else ","


def carica_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        prima = f.readline()
        f.seek(0)
        delim = rileva_delimitatore(prima)
        reader = csv.DictReader(f, delimiter=delim)
        fn = reader.fieldnames
        if not fn:
            raise ValueError(
                f"Il file '{path}' è vuoto o senza intestazione: impossibile "
                f"leggerlo. Il file potrebbe essere stato troncato da una "
                f"esecuzione interrotta; ripristinalo da un backup.")
        fieldnames = [h.strip() for h in fn]
        righe = [{(k.strip() if isinstance(k, str) else k): (v if isinstance(v, str) else "")
                 for k, v in row.items() if k is not None}
                for row in reader]
    return righe, fieldnames


def salva_csv(path, righe, fieldnames):
    """Salvataggio ATOMICO: scrive in un file temporaneo e poi lo sposta
    sopra l'originale con os.replace(), così un'interruzione durante la
    scrittura non corrompe il file originale."""
    path = Path(path)
    dirn = path.parent
    fd, tmp = tempfile.mkstemp(dir=dirn, prefix="." + path.name + ".", suffix=".tmp")
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(righe)
            f.flush()
            os.fsync(f.fileno())
        if os.path.getsize(tmp) == 0:
            raise RuntimeError("scrittura vuota: annullo per sicurezza")
        os.replace(tmp, path)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


def salva_xlsx_atomico(wb, out_path):
    """Salvataggio XLSX atomico: salva su file temporaneo poi os.replace."""
    out_path = Path(out_path)
    dirn = out_path.parent
    fd, tmp = tempfile.mkstemp(dir=dirn, prefix="." + out_path.name + ".", suffix=".tmp")
    os.close(fd)
    try:
        wb.save(tmp)
        if os.path.getsize(tmp) == 0:
            raise RuntimeError("scrittura vuota: annullo per sicurezza")
        os.replace(tmp, out_path)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


# ─── XLSX ──────────────────────────────────────────────────────────

def carica_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    fieldnames = [cell.value for cell in ws[1]]
    righe = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        righe.append({fieldnames[i]: (row[i] if i < len(row) else "") for i in range(len(fieldnames))})
    return righe, fieldnames, wb, ws


def salva_xlsx(ws, righe, fieldnames):
    for r_idx, riga in enumerate(righe, start=2):
        for c_idx, col in enumerate(fieldnames, start=1):
            ws.cell(row=r_idx, column=c_idx, value=riga.get(col, ""))


# ─── XLS ───────────────────────────────────────────────────────────

def carica_xls(path):
    import xlrd
    wb_xls = xlrd.open_workbook(path)
    ws = wb_xls.sheet_by_index(0)
    fieldnames = [str(cell.value).strip() for cell in ws.row(0)]
    righe = []
    for r in range(1, ws.nrows):
        righe.append({fieldnames[c]: str(ws.cell_value(r, c)) for c in range(len(fieldnames))})
    return righe, fieldnames


# ─── Core ─────────────────────────────────────────────────────────

def colonne_lingua(fieldnames):
    return [c for c in fieldnames if c not in ESCLUSE]


def scegli_modollo(modello_richiesto, modello_default=MODEL):
    if modello_richiesto:
        return modello_richiesto
    try:
        resp = requests.get(OLLAMA_TAGS_URL, timeout=10)
        resp.raise_for_status()
        modelli = resp.json().get("models", [])
    except Exception as e:
        log(f"Impossibile contattare Ollama per listare i modelli: {e}")
        log(f"Uso modello predefinito: {modello_default}")
        return modello_default

    if not modelli:
        log(f"Nessun modello trovato in Ollama. Uso: {modello_default}")
        return modello_default

    nomi = [m["name"] for m in modelli]

    if len(modelli) == 1:
        log(f"Modello disponibile: {nomi[0]}")
        return nomi[0]

    raccomandato = modello_default if modello_default in nomi else nomi[0]

    log("Modelli disponibili in Ollama:")
    for idx, m in enumerate(modelli, 1):
        tag = m.get("name", "?")
        size = m.get("size", 0)
        size_gb = size / 1e9
        consigliato = "  ← consigliato" if tag == raccomandato else ""
        log(f"  {idx}. {tag} ({size_gb:.1f}GB){consigliato}")

    while True:
        try:
            scelta = input(f"\nScegli modello [1-{len(modelli)}] (default: {raccomandato}): ").strip()
            if not scelta:
                return raccomandato
            idx = int(scelta) - 1
            if 0 <= idx < len(modelli):
                return nomi[idx]
        except (ValueError, IndexError):
            pass
        log("Scelta non valida. Riprova.")


def main():
    parser = argparse.ArgumentParser(description="Traduci CSV/XLSX/XLS con Ollama")
    parser.add_argument("file", help="Percorso del file da tradurre")
    parser.add_argument("--source-col", default=COLONNA_SOURCE,
                        help=f"Colonna sorgente (default: {COLONNA_SOURCE})")
    parser.add_argument("--save-interval", type=int, default=0,
                        help="Salva ogni N righe (0 = ad ogni riga tradotta, default: 0)")
    parser.add_argument("--delay", type=float, default=DELAY,
                        help=f"Secondi tra richieste (default: {DELAY})")
    parser.add_argument("-n", "--max-lines", type=int, default=0,
                        help="Traduci solo N righe (0=tutte, default: 0)")
    parser.add_argument("--model", default="",
                        help=f"Modello Ollama (default: {MODEL}; se omesso, lista interattiva)")
    parser.add_argument("--ollama-url", default=None,
                        help=f"URL base del server Ollama principale "
                             f"(default: {OLLAMA_BASE})")
    parser.add_argument("--ollama-url2", default=None,
                        help="URL base di un SECONDO server Ollama: il file "
                             "viene elaborato in parallelo (server1 in avanti "
                             "dalla metà iniziale, server2 all'indietro dalla "
                             "metà finale) per dimezzare il tempo.")
    parser.add_argument("--all", action="store_true",
                        help="Aggiungi tutte le lingue mancanti e traduci")
    parser.add_argument("--glossary", default=None,
                        help=f"File glossario multilingua (default: {GLOSSARIO_DEFAULT} se presente)")
    parser.add_argument("--no-glossary", action="store_true",
                        help="Disabilita l'uso del glossario")
    parser.add_argument("--glossary-overwrite", action="store_true",
                        help="Sovrascrivi anche le traduzioni già presenti usando il glossario")
    parser.add_argument("--rebuild-glossary", action="store_true",
                        help="Ricostruisci da zero la cache del glossario")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        log(f"File non trovato: {path}")
        sys.exit(1)
    if path.is_file() and path.stat().st_size == 0:
        log(f"ERRORE: il file '{path}' è vuoto (0 byte). Probabilmente è stato "
            f"troncato da un'esecuzione precedente interrotta. Ripristinalo "
            f"da un backup o scaricalo di nuovo.")
        sys.exit(1)

    col_source = args.source_col
    save_interval = args.save_interval
    if save_interval <= 0:
        save_interval = 1
    delay = args.delay
    modello = scegli_modollo(args.model)

    ext = path.suffix.lower()
    is_xlsx = ext in (".xlsx", ".xlsm")
    is_xls = ext == ".xls"

    wb = ws = out_path = None

    if is_xlsx:
        righe, fieldnames, wb, ws = carica_xlsx(path)
        out_path = path
        log(f"Caricato XLSX: {path.name}")
    elif is_xls:
        righe, fieldnames = carica_xls(path)
        out_path = path.with_suffix(".xlsx")
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(fieldnames)
        log(f"Caricato XLS: {path.name} -> salverò come {out_path.name}")
    else:
        righe, fieldnames = carica_csv(path)
        log(f"Caricato CSV: {path.name}")

    lingue = colonne_lingua(fieldnames)

    # Colonna sorgente mancante: fallback automatico invece di abortire.
    if col_source not in fieldnames:
        candidati = [col_source.lower(), "it", "source_term"]
        trovato = next((c for c in candidati if c in fieldnames), None)
        if trovato:
            log(f"Colonna sorgente '{col_source}' non trovata: "
                f"uso '{trovato}' come sorgente.")
            col_source = trovato
        else:
            log(f"Colonna sorgente '{col_source}' non trovata e nessun fallback "
                f"disponibile: per ogni lingua userò le altre traduzioni della "
                f"stessa riga come riferimento.")
            col_source = None

    # Le lingue da tradurre escludono la sorgente e la colonna 'source_term'
    lingue = [c for c in lingue if c != col_source and c != "source_term"]

    if args.all:
        for codice in TUTTE_LINGUE:
            if codice not in fieldnames:
                fieldnames.append(codice)
                for riga in righe:
                    riga[codice] = ""
        if wb is not None:
            for c_idx, col in enumerate(fieldnames, start=1):
                ws.cell(row=1, column=c_idx, value=col)
        lingue = [c for c in TUTTE_LINGUE if c != col_source and c != "source_term"]
        log(f"Aggiunte tutte le {len(lingue)} lingue disponibili")

    if not lingue:
        log("Nessuna colonna lingua trovata.")
        sys.exit(1)

    max_lines = args.max_lines
    da_elaborare = righe[:max_lines] if max_lines > 0 and max_lines < len(righe) else righe
    if max_lines > 0 and max_lines < len(righe):
        log(f"Tradurrò solo le prime {max_lines} righe su {len(righe)}")

    # ─── Glossario ───────────────────────────────────────────────
    glossario_path = None
    indice = {}
    gloss_cols = {}
    gloss_rows = {}

    if not args.no_glossary:
        gp = args.glossary
        if gp is None and Path(GLOSSARIO_DEFAULT).exists():
            gp = GLOSSARIO_DEFAULT
        if gp:
            if not Path(gp).exists():
                log(f"File glossario '{gp}' non trovato: proseguo senza glossario.")
            else:
                log(f"Carico glossario: {gp}")
                indice, gloss_cols = carica_glossario(gp, lingue, args.rebuild_glossary)
                # Pre-carica le righe del glossario utili per tutti i termini
                chiavi = set()
                for riga in da_elaborare:
                    s = norm(riga.get(col_source, ""))
                    if s:
                        chiavi.add(s)
                    for code in gloss_cols:
                        v = norm(riga.get(code, ""))
                        if v:
                            chiavi.add(v)
                record_ids = {indice[k] for k in chiavi if k in indice}
                gloss_rows = estrai_righe_glossario(gp, record_ids, gloss_cols)
                log(f"Glossario: {len(gloss_cols)} lingue mappate, "
                    f"{len(record_ids)} record rilevanti trovati.")
        else:
            log("Nessun glossario specificato: procedo senza.")
    else:
        log("Glossario disabilitato.")

    max_lines = args.max_lines
    da_elaborare = righe[:max_lines] if max_lines > 0 and max_lines < len(righe) else righe
    if max_lines > 0 and max_lines < len(righe):
        log(f"Tradurrò solo le prime {max_lines} righe su {len(righe)}")

    log(f"Colonna sorgente: {col_source}")
    log(f"Lingue destinazione: {', '.join(lingue)}")
    log(f"Righe: {len(righe)}")
    log("")

    session = requests.Session()

    esc_thread = avvia_controllo_esc()
    log("Puoi interrompere in qualsiasi momento premendo ESC "
        "(il progresso verrà salvato).")
    log("")

    def process_row(i, riga, session, base_url, ritmo):
        """Traduce una singola riga (muta `riga`). Restituisce il numero di
        lingue tradotte. Sicuro da chiamare da più thread (usa solo letture
        sulle strutture condivise e scrive solo su `riga`)."""
        if col_source is not None:
            termine = riga.get(col_source, "")
            if not termine.strip():
                log(f"[{i}/{len(da_elaborare)}] (riga vuota, saltata)")
                return 0
            sorgente_riga = termine
        else:
            sorgente_riga = None  # calcolata per ogni lingua

        # Trova i record del glossario corrispondenti a questa riga:
        # usando il termine sorgente e le traduzioni già presenti.
        matching = []
        if col_source is not None:
            s = norm(sorgente_riga)
            if s and s in indice:
                matching.append(indice[s])
        for code in gloss_cols:
            v = norm(riga.get(code, ""))
            if v and v in indice:
                matching.append(indice[v])
        matching = list(dict.fromkeys(matching))  # unici, ordine preservato

        tradotte = 0
        for lingua in lingue:
            if interruzione.is_set():
                break

            # Sorgente: la colonna sorgente, oppure (se assente) la prima
            # altra lingua non vuota sulla riga.
            if col_source is not None:
                sorg = sorgente_riga
            else:
                sorg = ""
                for c in lingue:
                    if c == lingua:
                        continue
                    v = riga.get(c, "").strip()
                    if v:
                        sorg = v
                        break
                if not sorg:
                    continue

            gcol = gloss_cols.get(lingua)
            val_attuale = riga.get(lingua, "")
            if val_attuale.strip() and not args.glossary_overwrite:
                continue

            # 1) Prova a prendere la traduzione dal glossario
            gval = ""
            if gcol:
                for rid in matching:
                    d = gloss_rows.get(rid)
                    if d and d.get(gcol):
                        gval = prima_variante(d[gcol])
                        break

            if gval:
                riga[lingua] = gval
                log(f"{colore_lingua(lingua)}: {gval}  (glossario)")
                tradotte += 1
                continue

            # 2) Altrimenti usa l'LLM, ma passando come riferimento le
            #    altre traduzioni già presenti sulla riga (o nel glossario)
            #    per migliorare coerenza e completezza.
            contesto = {}
            for code in gloss_cols:
                if code == lingua:
                    continue
                v = riga.get(code, "").strip()
                if not v:
                    for rid in matching:
                        d = gloss_rows.get(rid)
                        if d and d.get(gloss_cols[code]):
                            v = prima_variante(d[gloss_cols[code]])
                            break
                if v:
                    contesto[code] = v

            trad = traduci(sorg, lingua, modello=modello,
                           session=session, contesto=contesto,
                           base_url=base_url, ritmo=ritmo)
            riga[lingua] = trad
            log(f"{colore_lingua(lingua)}: {trad}")
            tradotte += 1
            if ritmo.delay > 0:
                time.sleep(ritmo.delay)

        intestazione = termine if col_source is not None else (
            riga.get("source_term", "") or "(multi-sorgente)")
        if tradotte == 0 and not interruzione.is_set():
            log(f"[{i}/{len(da_elaborare)}] '{intestazione}' — già completo")
        else:
            log(f"[{i}/{len(da_elaborare)}] '{intestazione}' — {tradotte} lingue tradotte")
        return tradotte

    def salva_ora():
        if wb is not None:
            salva_xlsx(ws, righe, fieldnames)
            salva_xlsx_atomico(wb, out_path)
        else:
            salva_csv(path, righe, fieldnames)

    lavori = list(enumerate(da_elaborare, 1))  # [(i, riga), ...]
    base1 = normalizza_base(args.ollama_url or OLLAMA_BASE)
    base2 = normalizza_base(args.ollama_url2) if args.ollama_url2 else None
    url1 = base1 + "/api/generate"
    url2 = base2 + "/api/generate" if base2 else None

    try:
        if url2:
            # ─── Modalità parallela su due server ──────────────────────
            # server1 elabora la prima metà in avanti; server2 elabora la
            # seconda metà a ritroso (dalla fine), così si incontrano a metà.
            # Ogni server ha il suo ritmo (delay adattivo) indipendente.
            log(f"Modalità PARALLELA: server1={base1} (avanti), "
                f"server2={base2} (indietro dalla fine).")
            mid = len(lavori) // 2
            l1 = lavori[:mid]
            l2 = list(reversed(lavori[mid:]))  # dalla fine
            s1 = requests.Session()
            s2 = requests.Session()
            r1 = Ritmo(args.delay or DELAY)
            r2 = Ritmo(args.delay or DELAY)

            def worker(sub, sess, url, ritmo):
                for i, riga in sub:
                    if interruzione.is_set():
                        break
                    process_row(i, riga, sess, url, ritmo)

            t1 = threading.Thread(target=worker, args=(l1, s1, url1, r1), daemon=True)
            t2 = threading.Thread(target=worker, args=(l2, s2, url2, r2), daemon=True)
            t1.start()
            t2.start()
            while t1.is_alive() or t2.is_alive():
                t1.join(0.2)
                t2.join(0.2)
                if interruzione.is_set():
                    break
            interruzione.set()  # ferma i worker se non già fermi
            t1.join()
            t2.join()
            salva_ora()
            log("Elaborazione parallela completata e salvata.")
        else:
            # ─── Modalità singolo server (comportamento standard) ───────
            ritmo = Ritmo(args.delay or DELAY)
            for i, riga in lavori:
                if interruzione.is_set():
                    log("Interruzione (ESC) richiesta: salvo ed esco.")
                    break
                process_row(i, riga, session, url1, ritmo)
                if wb is not None:
                    if save_interval > 0 and i % save_interval == 0:
                        salva_xlsx(ws, righe, fieldnames)
                        salva_xlsx_atomico(wb, out_path)
                else:
                    if save_interval > 0 and i % save_interval == 0:
                        salva_csv(path, righe, fieldnames)
                log("")
    except KeyboardInterrupt:
        log("\nInterruzione (Ctrl+C) richiesta: salvo ed esco.")
    finally:
        salva_ora()
        ferma_controllo_esc(esc_thread)

    log("Fatto! Tutte le traduzioni completate/salvate.")


if __name__ == "__main__":
    main()
